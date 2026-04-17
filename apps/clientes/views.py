from django.shortcuts import render, get_object_or_404
from django.contrib.auth.decorators import login_required
from .models import Cliente
from apps.clientes.services import obtener_turno_actual, calcular_estado_cliente
from django.db.models import Q
from .forms import ClienteForm
from django.http import HttpResponse, HttpResponseForbidden


@login_required
def dashboard(request):
    """
    Vista principal del dashboard.
    Si es request HTMX, devuelve solo el partial.
    Si es request normal, devuelve página completa.
    """
    # Obtener todos los turnos activos
    from apps.usuarios.models import Turno

    turnos = Turno.objects.filter(activo=True).order_by("hora_inicio")

    # Determinar qué turno mostrar
    turno_seleccionado_id = request.GET.get("turno")

    if turno_seleccionado_id:
        # Usuario seleccionó un turno manualmente
        turno_actual = Turno.objects.filter(
            id=turno_seleccionado_id, activo=True
        ).first()
    else:
        # Detectar turno automáticamente según la hora
        turno_actual = obtener_turno_actual()

    # Obtener parámetros de búsqueda y filtro
    busqueda = request.GET.get("busqueda", "").strip()
    filtro_estado = request.GET.get("estado", "")

    # Query base: clientes del turno actual
    clientes_qs = Cliente.objects.filter(activo=True)

    # Filtrar por turno
    if turno_actual and not busqueda:
        clientes_qs = clientes_qs.filter(turno=turno_actual)

    # Aplicar búsqueda (nombre O apellido)
    if busqueda:
        clientes_qs = clientes_qs.filter(
            Q(nombre__icontains=busqueda) | Q(apellido__icontains=busqueda)
        )

    # Optimización: traer relaciones en una query
    clientes_qs = clientes_qs.select_related("turno", "usuario_creador")

    # Calcular estado y aplicar filtro de estado
    clientes = []
    for cliente in clientes_qs:
        cliente.estado_actual = calcular_estado_cliente(cliente)

        # Filtrar por estado si se seleccionó uno
        if filtro_estado:
            if filtro_estado == "al_dia" and cliente.estado_actual != "al_dia":
                continue
            elif filtro_estado == "vencido" and cliente.estado_actual != "vencido":
                continue
            elif (
                filtro_estado == "pendiente"
                and cliente.estado_actual != "pendiente_consulta"
            ):
                continue

        clientes.append(cliente)

    context = {
        "turnos": turnos,
        "turno_actual": turno_actual,
        "clientes": clientes,
        "busqueda": busqueda,
        "filtro_estado": filtro_estado,
    }

    # Si es request HTMX, devolver solo el partial
    if request.headers.get("HX-Request"):
        return render(request, "clientes/_tabla_clientes.html", context)

    # Si es request normal, devolver página completa
    return render(request, "clientes/dashboard.html", context)


@login_required
def modal_historial_pagos(request, cliente_id):
    """
    Muestra el historial de pagos de un cliente.
    """
    cliente = get_object_or_404(Cliente, id=cliente_id)

    # Obtener pagos ordenados por más reciente
    pagos = cliente.pagos.all().order_by("-fecha_pago")

    context = {
        "cliente": cliente,
        "pagos": pagos,
    }

    return render(request, "clientes/_modal_historial.html", context)


@login_required
def crear_cliente(request):
    if request.method == "POST":
        form = ClienteForm(request.POST)
        if form.is_valid():
            cliente = form.save(commit=False)
            cliente.usuario_creador = request.user
            cliente.save()
            return HttpResponse(
                status=204, headers={"HX-Trigger": "clienteActualizado"}
            )
    else:
        form = ClienteForm()
    return render(request, "clientes/_modal_cliente.html", {"form": form})


@login_required
def editar_cliente(request, cliente_id):
    cliente = get_object_or_404(Cliente, id=cliente_id)
    if request.method == "POST":
        form = ClienteForm(request.POST, instance=cliente)
        if form.is_valid():
            form.save()
            return HttpResponse(
                status=204, headers={"HX-Trigger": "clienteActualizado"}
            )
    else:
        form = ClienteForm(instance=cliente)
    return render(
        request, "clientes/_modal_cliente.html", {"form": form, "cliente": cliente}
    )


@login_required
def reportes(request):
    if request.user.rol != "dueño":
        return HttpResponseForbidden("No tenés permiso para ver esta página.")

    from apps.usuarios.models import Turno
    from apps.pagos.models import Pago
    from django.db.models import Sum, Count
    from datetime import date

    hoy = date.today()
    mes_actual = hoy.replace(day=1)
    turnos = Turno.objects.filter(activo=True).order_by("hora_inicio")

    datos_turnos = []
    for turno in turnos:
        clientes_turno = Cliente.objects.filter(
            turno=turno, activo=True
        ).select_related("plan")

        total = clientes_turno.count()
        al_dia = vencidos = pendientes = 0

        for cliente in clientes_turno:
            estado = calcular_estado_cliente(cliente, hoy)
            if estado == "al_dia":
                al_dia += 1
            elif estado == "vencido":
                vencidos += 1
            elif estado == "pendiente_consulta":
                pendientes += 1

        ingresos_mes = (
            Pago.objects.filter(
                cliente__turno=turno,
                mes_cubierto=mes_actual,
            ).aggregate(total=Sum("monto"))["total"]
            or 0
        )

        porcentaje_vencidos = round((vencidos / total * 100), 1) if total > 0 else 0

        datos_turnos.append(
            {
                "turno": turno,
                "total": total,
                "al_dia": al_dia,
                "vencidos": vencidos,
                "pendientes": pendientes,
                "porcentaje_vencidos": porcentaje_vencidos,
                "ingresos_mes": ingresos_mes,
            }
        )

    return render(
        request,
        "clientes/reportes.html",
        {
            "datos_turnos": datos_turnos,
            "mes_actual": mes_actual,
        },
    )


# *******************************
# VISTAS EXPORT/IMPORT DATOS XML
# *******************************


@login_required
def exportar_xml(request):
    if request.user.rol != "dueño":
        return HttpResponseForbidden()

    from django.http import HttpResponse
    from apps.pagos.models import Pago
    import xml.etree.ElementTree as ET

    root = ET.Element("gimnasio")

    # Clientes
    clientes_el = ET.SubElement(root, "clientes")
    for cliente in Cliente.objects.select_related("plan", "turno").all():
        c = ET.SubElement(clientes_el, "cliente")
        ET.SubElement(c, "id").text = str(cliente.id)
        ET.SubElement(c, "nombre").text = cliente.nombre
        ET.SubElement(c, "apellido").text = cliente.apellido
        ET.SubElement(c, "telefono").text = cliente.telefono or ""
        ET.SubElement(c, "email").text = cliente.email or ""
        ET.SubElement(c, "plan").text = cliente.plan.codigo
        ET.SubElement(c, "turno").text = cliente.turno.nombre
        ET.SubElement(c, "activo").text = str(cliente.activo)

    # Pagos
    pagos_el = ET.SubElement(root, "pagos")
    for pago in Pago.objects.select_related("cliente").all():
        p = ET.SubElement(pagos_el, "pago")
        ET.SubElement(p, "id").text = str(pago.id)
        ET.SubElement(p, "cliente_id").text = str(pago.cliente.id)
        ET.SubElement(p, "fecha_pago").text = str(pago.fecha_pago)
        ET.SubElement(p, "mes_cubierto").text = str(pago.mes_cubierto)
        ET.SubElement(p, "monto").text = str(pago.monto)
        ET.SubElement(p, "observaciones").text = pago.observaciones or ""

    tree = ET.ElementTree(root)
    ET.indent(tree, space="  ")

    response = HttpResponse(content_type="application/xml")
    response["Content-Disposition"] = 'attachment; filename="gimnasio_backup.xml"'
    tree.write(response, encoding="unicode", xml_declaration=True)
    return response


@login_required
def importar_xml(request):
    if request.user.rol != "dueño":
        return HttpResponseForbidden()

    if request.method != "POST":
        return render(request, "clientes/_importar_xml.html")

    import xml.etree.ElementTree as ET
    from apps.pagos.models import Pago
    from datetime import date

    archivo = request.FILES.get("archivo")
    if not archivo:
        return render(
            request,
            "clientes/_importar_xml.html",
            {"error": "No se seleccionó archivo"},
        )

    try:
        tree = ET.parse(archivo)
        root = tree.getroot()
    except ET.ParseError:
        return render(
            request,
            "clientes/_importar_xml.html",
            {"error": "El archivo no es un XML válido"},
        )

    from apps.usuarios.models import Turno
    from apps.clientes.models import Plan

    clientes_creados = 0
    clientes_saltados = 0
    pagos_creados = 0
    pagos_saltados = 0

    # Mapa id_xml → cliente nuevo (para los pagos)
    mapa_clientes = {}

    for c in root.findall("clientes/cliente"):
        id_xml = c.findtext("id")
        nombre = c.findtext("nombre")
        apellido = c.findtext("apellido")

        # Buscar plan y turno por codigo/nombre
        plan = Plan.objects.filter(codigo=c.findtext("plan")).first()
        turno = Turno.objects.filter(nombre=c.findtext("turno")).first()

        if not plan or not turno:
            clientes_saltados += 1
            continue

        # Buscar si ya existe un cliente con mismo nombre, apellido y turno
        existente = Cliente.objects.filter(
            nombre=nombre, apellido=apellido, turno=turno
        ).first()

        if existente:
            mapa_clientes[id_xml] = existente
            clientes_saltados += 1
            continue

        cliente = Cliente.objects.create(
            nombre=nombre,
            apellido=apellido,
            telefono=c.findtext("telefono") or "",
            email=c.findtext("email") or "",
            plan=plan,
            turno=turno,
            activo=c.findtext("activo") == "True",
            usuario_creador=request.user,
        )
        mapa_clientes[id_xml] = cliente
        clientes_creados += 1

    for p in root.findall("pagos/pago"):
        cliente_id_xml = p.findtext("cliente_id")
        cliente = mapa_clientes.get(cliente_id_xml)

        if not cliente:
            pagos_saltados += 1
            continue

        mes_cubierto_str = p.findtext("mes_cubierto")
        fecha_pago_str = p.findtext("fecha_pago")

        try:
            mes_cubierto = date.fromisoformat(mes_cubierto_str)
            # Normalizar a día 1
            mes_cubierto = mes_cubierto.replace(day=1)
            fecha_pago = date.fromisoformat(fecha_pago_str)
        except (ValueError, TypeError):
            pagos_saltados += 1
            continue

        # Saltar si ya existe pago para ese cliente y mes
        if Pago.objects.filter(cliente=cliente, mes_cubierto=mes_cubierto).exists():
            pagos_saltados += 1
            continue

        Pago.objects.create(
            cliente=cliente,
            fecha_pago=fecha_pago,
            mes_cubierto=mes_cubierto,
            monto=p.findtext("monto") or 0,
            observaciones=p.findtext("observaciones") or "",
            usuario_registrador=request.user,
        )
        pagos_creados += 1

    return render(
        request,
        "clientes/_importar_xml.html",
        {
            "resultado": {
                "clientes_creados": clientes_creados,
                "clientes_saltados": clientes_saltados,
                "pagos_creados": pagos_creados,
                "pagos_saltados": pagos_saltados,
            }
        },
    )


# *******************************
# VISTAS EXPORT/IMPORT DATOS EXCEL
# *******************************


@login_required
def exportar_excel(request):
    if request.user.rol != "dueño":
        return HttpResponseForbidden()

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from apps.pagos.models import Pago
    from django.http import HttpResponse

    wb = Workbook()

    # ── Hoja 1: Clientes ──────────────────────────────────────────────────────
    ws_clientes = wb.active
    ws_clientes.title = "Clientes"

    headers_clientes = [
        "ID",
        "Apellido",
        "Nombre",
        "Plan",
        "Turno",
        "Teléfono",
        "Email",
        "Activo",
    ]
    ws_clientes.append(headers_clientes)

    # Estilo encabezado
    for col in range(1, len(headers_clientes) + 1):
        cell = ws_clientes.cell(row=1, column=col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2563EB")
        cell.alignment = Alignment(horizontal="center")

    for cliente in Cliente.objects.select_related("plan", "turno").all():
        ws_clientes.append(
            [
                cliente.id,
                cliente.apellido,
                cliente.nombre,
                cliente.plan.codigo,
                cliente.turno.nombre,
                cliente.telefono or "",
                cliente.email or "",
                "Sí" if cliente.activo else "No",
            ]
        )

    # Ancho de columnas automático
    for col in ws_clientes.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws_clientes.column_dimensions[col[0].column_letter].width = max_len + 4

    # ── Hoja 2: Pagos ─────────────────────────────────────────────────────────
    ws_pagos = wb.create_sheet("Pagos")

    headers_pagos = [
        "ID",
        "Cliente ID",
        "Apellido",
        "Nombre",
        "Fecha Pago",
        "Mes Cubierto",
        "Monto",
        "Observaciones",
    ]
    ws_pagos.append(headers_pagos)

    for col in range(1, len(headers_pagos) + 1):
        cell = ws_pagos.cell(row=1, column=col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2563EB")
        cell.alignment = Alignment(horizontal="center")

    for pago in Pago.objects.select_related("cliente").all():
        ws_pagos.append(
            [
                pago.id,
                pago.cliente.id,
                pago.cliente.apellido,
                pago.cliente.nombre,
                pago.fecha_pago.strftime("%d/%m/%Y"),
                pago.mes_cubierto.strftime("%m/%Y"),
                float(pago.monto),
                pago.observaciones or "",
            ]
        )

    for col in ws_pagos.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws_pagos.column_dimensions[col[0].column_letter].width = max_len + 4

    # ── Respuesta ─────────────────────────────────────────────────────────────
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="gimnasio_backup.xlsx"'
    wb.save(response)
    return response


@login_required
def importar_excel(request):
    if request.user.rol != "dueño":
        return HttpResponseForbidden()

    if request.method != "POST":
        return render(request, "clientes/_importar_excel.html")

    from openpyxl import load_workbook
    from apps.pagos.models import Pago
    from apps.usuarios.models import Turno
    from apps.clientes.models import Plan
    from datetime import date

    archivo = request.FILES.get("archivo")
    if not archivo:
        return render(
            request,
            "clientes/_importar_excel.html",
            {"error": "No se seleccionó archivo"},
        )

    try:
        wb = load_workbook(archivo)
    except Exception:
        return render(
            request,
            "clientes/_importar_excel.html",
            {"error": "El archivo no es un Excel válido"},
        )

    clientes_creados = 0
    clientes_saltados = 0
    pagos_creados = 0
    pagos_saltados = 0
    errores = []

    # ── Hoja Clientes ─────────────────────────────────────────────────────────
    if "Clientes" not in wb.sheetnames:
        return render(
            request,
            "clientes/_importar_excel.html",
            {"error": "El archivo no tiene hoja 'Clientes'"},
        )

    ws_clientes = wb["Clientes"]
    mapa_clientes = {}

    for i, row in enumerate(
        ws_clientes.iter_rows(min_row=2, values_only=True), start=2
    ):
        if not any(row):
            continue

        (
            id_excel,
            apellido,
            nombre,
            plan_codigo,
            turno_nombre,
            telefono,
            email,
            activo,
        ) = row

        plan = Plan.objects.filter(codigo=plan_codigo).first()
        turno = Turno.objects.filter(nombre=turno_nombre).first()

        if not plan:
            errores.append(f"Fila {i} (Clientes): plan '{plan_codigo}' no existe")
            clientes_saltados += 1
            continue

        if not turno:
            errores.append(f"Fila {i} (Clientes): turno '{turno_nombre}' no existe")
            clientes_saltados += 1
            continue

        existente = Cliente.objects.filter(
            nombre=nombre, apellido=apellido, turno=turno
        ).first()

        if existente:
            mapa_clientes[id_excel] = existente
            clientes_saltados += 1
            continue

        cliente = Cliente.objects.create(
            nombre=nombre,
            apellido=apellido,
            telefono=str(telefono) if telefono else "",
            email=email or "",
            plan=plan,
            turno=turno,
            activo=activo == "Sí",
            usuario_creador=request.user,
        )
        mapa_clientes[id_excel] = cliente
        clientes_creados += 1

    # ── Hoja Pagos ────────────────────────────────────────────────────────────
    if "Pagos" in wb.sheetnames:
        ws_pagos = wb["Pagos"]

        for i, row in enumerate(
            ws_pagos.iter_rows(min_row=2, values_only=True), start=2
        ):
            if not any(row):
                continue

            (
                id_excel,
                cliente_id_excel,
                apellido,
                nombre,
                fecha_pago_str,
                mes_cubierto_str,
                monto,
                observaciones,
            ) = row

            cliente = mapa_clientes.get(cliente_id_excel)
            if not cliente:
                pagos_saltados += 1
                continue

            try:
                if isinstance(fecha_pago_str, str):
                    from datetime import datetime

                    fecha_pago = datetime.strptime(fecha_pago_str, "%d/%m/%Y").date()
                else:
                    fecha_pago = fecha_pago_str

                if isinstance(mes_cubierto_str, str):
                    mes_cubierto = (
                        datetime.strptime(mes_cubierto_str, "%m/%Y")
                        .date()
                        .replace(day=1)
                    )
                else:
                    mes_cubierto = mes_cubierto_str.replace(day=1)
            except (ValueError, TypeError, AttributeError):
                errores.append(f"Fila {i} (Pagos): fecha inválida")
                pagos_saltados += 1
                continue

            if Pago.objects.filter(cliente=cliente, mes_cubierto=mes_cubierto).exists():
                pagos_saltados += 1
                continue

            Pago.objects.create(
                cliente=cliente,
                fecha_pago=fecha_pago,
                mes_cubierto=mes_cubierto,
                monto=monto or 0,
                observaciones=observaciones or "",
                usuario_registrador=request.user,
            )
            pagos_creados += 1

    return render(
        request,
        "clientes/_importar_excel.html",
        {
            "resultado": {
                "clientes_creados": clientes_creados,
                "clientes_saltados": clientes_saltados,
                "pagos_creados": pagos_creados,
                "pagos_saltados": pagos_saltados,
            },
            "errores": errores,
        },
    )


# *********************************
# VISTAS EXPORT/IMPORT DATOS A CSV
# *********************************


@login_required
def exportar_csv(request):
    if request.user.rol != "dueño":
        return HttpResponseForbidden()

    import csv
    from apps.pagos.models import Pago
    from django.http import HttpResponse

    response = HttpResponse(content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = 'attachment; filename="gimnasio_backup.csv"'
    response.write("\ufeff")  # BOM para que Excel lo abra bien

    writer = csv.writer(response)

    # Clientes
    writer.writerow(["## CLIENTES"])
    writer.writerow(
        ["ID", "Apellido", "Nombre", "Plan", "Turno", "Teléfono", "Email", "Activo"]
    )
    for cliente in Cliente.objects.select_related("plan", "turno").all():
        writer.writerow(
            [
                cliente.id,
                cliente.apellido,
                cliente.nombre,
                cliente.plan.codigo,
                cliente.turno.nombre,
                cliente.telefono or "",
                cliente.email or "",
                "Sí" if cliente.activo else "No",
            ]
        )

    # Separador
    writer.writerow([])
    writer.writerow(["## PAGOS"])
    writer.writerow(
        [
            "ID",
            "Cliente ID",
            "Apellido",
            "Nombre",
            "Fecha Pago",
            "Mes Cubierto",
            "Monto",
            "Observaciones",
        ]
    )
    for pago in Pago.objects.select_related("cliente").all():
        writer.writerow(
            [
                pago.id,
                pago.cliente.id,
                pago.cliente.apellido,
                pago.cliente.nombre,
                pago.fecha_pago.strftime("%d/%m/%Y"),
                pago.mes_cubierto.strftime("%m/%Y"),
                float(pago.monto),
                pago.observaciones or "",
            ]
        )

    return response


@login_required
def importar_csv(request):
    if request.user.rol != "dueño":
        return HttpResponseForbidden()

    if request.method != "POST":
        return render(request, "clientes/_importar_csv.html")

    import csv
    from apps.pagos.models import Pago
    from apps.usuarios.models import Turno
    from apps.clientes.models import Plan
    from datetime import datetime

    archivo = request.FILES.get("archivo")
    if not archivo:
        return render(
            request,
            "clientes/_importar_csv.html",
            {"error": "No se seleccionó archivo"},
        )

    try:
        contenido = archivo.read().decode("utf-8-sig")
        lines = contenido.splitlines()
        reader = csv.reader(lines)
    except Exception:
        return render(
            request,
            "clientes/_importar_csv.html",
            {"error": "El archivo no es un CSV válido"},
        )

    clientes_creados = 0
    clientes_saltados = 0
    pagos_creados = 0
    pagos_saltados = 0
    errores = []
    mapa_clientes = {}

    seccion = None

    for i, row in enumerate(reader, start=1):
        if not any(row):
            continue

        if row[0] == "## CLIENTES":
            seccion = "clientes"
            continue
        if row[0] == "## PAGOS":
            seccion = "pagos"
            continue
        if row[0] in ["ID", "Apellido"]:
            continue

        if seccion == "clientes":
            if len(row) < 8:
                errores.append(f"Fila {i}: datos incompletos")
                continue

            (
                id_csv,
                apellido,
                nombre,
                plan_codigo,
                turno_nombre,
                telefono,
                email,
                activo,
            ) = row[:8]

            plan = Plan.objects.filter(codigo=plan_codigo).first()
            turno = Turno.objects.filter(nombre=turno_nombre).first()

            if not plan:
                errores.append(f"Fila {i}: plan '{plan_codigo}' no existe")
                clientes_saltados += 1
                continue

            if not turno:
                errores.append(f"Fila {i}: turno '{turno_nombre}' no existe")
                clientes_saltados += 1
                continue

            existente = Cliente.objects.filter(
                nombre=nombre, apellido=apellido, turno=turno
            ).first()

            if existente:
                mapa_clientes[id_csv] = existente
                clientes_saltados += 1
                continue

            cliente = Cliente.objects.create(
                nombre=nombre,
                apellido=apellido,
                telefono=telefono or "",
                email=email or "",
                plan=plan,
                turno=turno,
                activo=activo == "Sí",
                usuario_creador=request.user,
            )
            mapa_clientes[id_csv] = cliente
            clientes_creados += 1

        elif seccion == "pagos":
            if len(row) < 8:
                errores.append(f"Fila {i}: datos incompletos")
                continue

            (
                id_csv,
                cliente_id_csv,
                apellido,
                nombre,
                fecha_pago_str,
                mes_cubierto_str,
                monto,
                observaciones,
            ) = row[:8]

            cliente = mapa_clientes.get(cliente_id_csv)
            if not cliente:
                pagos_saltados += 1
                continue

            try:
                fecha_pago = datetime.strptime(fecha_pago_str, "%d/%m/%Y").date()
                mes_cubierto = (
                    datetime.strptime(mes_cubierto_str, "%m/%Y").date().replace(day=1)
                )
            except ValueError:
                errores.append(f"Fila {i}: fecha inválida")
                pagos_saltados += 1
                continue

            if Pago.objects.filter(cliente=cliente, mes_cubierto=mes_cubierto).exists():
                pagos_saltados += 1
                continue

            Pago.objects.create(
                cliente=cliente,
                fecha_pago=fecha_pago,
                mes_cubierto=mes_cubierto,
                monto=monto or 0,
                observaciones=observaciones or "",
                usuario_registrador=request.user,
            )
            pagos_creados += 1

    return render(
        request,
        "clientes/_importar_csv.html",
        {
            "resultado": {
                "clientes_creados": clientes_creados,
                "clientes_saltados": clientes_saltados,
                "pagos_creados": pagos_creados,
                "pagos_saltados": pagos_saltados,
            },
            "errores": errores,
        },
    )
